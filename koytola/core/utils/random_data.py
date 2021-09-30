import os
import random
import string
import unicodedata
import uuid
from unittest.mock import patch
from user_agents import parse
from django.contrib.auth.hashers import make_password
from django.conf import settings
from django.contrib.auth.models import Group, Permission
from django.contrib.sites.models import Site
from django.core.files import File
from django.urls import reverse
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.text import slugify
from faker import Factory
from faker.providers import BaseProvider
from measurement.measures import Weight
from prices import Money
import openpyxl
from ...account import AccountEvents
from ...account.models import Address, User, AccountEvent, Countries
from ...account.utils import store_user_address, AddressType
from ...core.permissions import (
    AccountPermissions,
    ProfilePermissions,
    OrderPermissions,
    get_permissions,
)
from ...product.models import HSCodeAndProduct
from ...analytics import TrackingTypes
from ...analytics.models import Tracking
from ...core.utils import build_absolute_uri
from ...graphql.analytics.utils import (
    get_device_type,
    get_geo_details_from_ip
)
from ...menu.models import Menu
from ...news import NewsAudienceType
from ...news.models import News
from ...page.models import Page, PageType
from ...payment import gateway
from ...payment.utils import create_payment
from ...product import ProductUnits
from ...product.models import (
    Category,
    Product,
    ProductUnits,
    ProductVideo,
)
from ...profile import (
    CompanySize,
    CompanyType,
    EmployeeNumber,

)
from ...profile.models import (
    Company,
    Roetter,
    CertificateType,
    Industry
)
import json
from django.db.models import Q
import csv
import urllib
import io
from django.utils.text import slugify
from django.core.files.storage import FileSystemStorage
from xml.etree import ElementTree
import json
import xmltodict
import requests
from datetime import datetime

fake = Factory.create()
DUMMY_STAFF_PASSWORD = "password"
REFERRER_OPTIONS = [
    "https://instagram.com",
    "https://youtube.com",
    "https://facebook.com",
    "https://google.com",
    "https://tiktok.com",
    "https://snapchat.com",
    "https://clubhouse.com",
    "",
    "",
    "",
    ""
]


def _get_random_model_instance(
        model,
        is_active=None,
        is_published=None
):
    queryset = model.objects.all()
    if is_active and hasattr(model, 'is_active'):
        queryset = queryset.filter(is_active=True)
    if is_published and hasattr(model, 'is_published'):
        queryset = queryset.filter(is_published=True)
    return queryset.order_by('?')[0]


def set_field_as_money(defaults, field):
    amount_field = f"{field}_amount"
    if amount_field in defaults and defaults[amount_field] is not None:
        defaults[field] = Money(defaults[amount_field], settings.DEFAULT_CURRENCY)


class SiteProvider(BaseProvider):
    def money(self):
        return Money(fake.pydecimal(2, 2, positive=True), settings.DEFAULT_CURRENCY)

    def weight(self):
        return Weight(kg=fake.pydecimal(1, 2, positive=True))


fake.add_provider(SiteProvider)


def get_email(first_name, last_name):
    _first = unicodedata.normalize("NFD", first_name).encode("ascii", "ignore")
    _last = unicodedata.normalize("NFD", last_name).encode("ascii", "ignore")
    return "%s.%s@example.com" % (
        _first.lower().decode("utf-8"),
        _last.lower().decode("utf-8"),
    )


def create_address(save=True):
    address = Address(
        first_name=fake.first_name(),
        last_name=fake.last_name(),
        street_address_1=fake.street_address(),
        city=fake.city(),
        country=settings.DEFAULT_COUNTRY,
    )

    if address.country == "US":
        state = fake.state_abbr()
        address.country_area = state
        address.postal_code = fake.postalcode_in_state(state)
    else:
        address.postal_code = fake.postalcode()

    if save:
        address.save()
    return address


def create_fake_event(save=True):
    event_type = random.choice(AccountEvents.CHOICES)[0]
    event = AccountEvent(
        type=event_type.upper()
    )

    user_index = User.objects.accounts().order_by("?")[0].id
    if "company" in event_type:
        company = Company.objects.filter(user_id=user_index).first()
        if company:
            event.company_id = company.id
    event.user_id = user_index

    if save:
        event.save()
    return event


def create_fake_user(save=True):
    address = create_address(save=save)
    email = get_email(address.first_name, address.last_name)

    # Skip the email if it already exists
    try:
        return User.objects.get(email=email)
    except User.DoesNotExist:
        pass

    user = User(
        first_name=address.first_name,
        last_name=address.last_name,
        email=email,
        password="password",
        default_billing_address=address,
        default_shipping_address=address,
        is_active=True,
        note=fake.paragraph(),
        date_joined=fake.date_time(tzinfo=timezone.get_current_timezone()),
    )

    if save:
        user.save()
        user.addresses.add(address)
    return user


# We don't want to spam the console with payment confirmations sent to
# fake customers.
@patch("koytola.order.emails.send_payment_confirmation.delay")
def create_fake_payment(mock_email_confirmation, order):
    payment = create_payment(
        gateway="koytola..payments.dummy",
        customer_ip_address=fake.ipv4(),
        email=order.user_email,
        order=order,
        payment_token=str(uuid.uuid4()),
        total=order.total.gross.amount,
        currency=order.total.gross.currency,
    )

    # Create authorization transaction
    gateway.authorize(payment, payment.token)
    # 20% chance to void the transaction at this stage
    if random.choice([0, 0, 0, 0, 1]):
        gateway.void(payment)
        return payment
    # 25% to end the payment at the authorization stage
    if not random.choice([1, 1, 1, 0]):
        return payment
    # Create capture transaction
    gateway.capture(payment)
    # 25% to refund the payment
    if random.choice([0, 0, 0, 1]):
        gateway.refund(payment)
    return payment


def create_fake_company(save=True, user_index=None, user=None):
    name = fake.company()
    slug = slugify(name, allow_unicode=True)
    slug_check = True
    while slug_check:
        if Company.objects.filter(slug=slug):
            slug = slug + get_random_string()
        else:
            slug_check = False
    website = "https://" + name.replace(" ", "").lower() + ".com"
    company = Company(
        name=name,
        slug=slug,
        website=website,
        # email=fake.company_email(),
        phone=fake.phone_number(),
        content={
            "blocks": [
                {
                    "data": {"text": fake.catch_phrase()},
                    "type": "paragraph",
                },
            ]
        },
        is_published=True if random.random() <= 0.98 else False,
    )

    if user:
        company.address = user.addresses.first()
        company.user_id = user.id
    else:
        if user_index and type(user_index) == int:
            company.user_id = user_index
            user = User.objects.filter(id=user_index).first()
            company.address = user.addresses.first()

    if random.random() < 0.5:
        company.no_of_employees = random.choice(EmployeeNumber.CHOICES)[0]
    if random.random() < 0.5:
        company.founded_year = random.randint(1900, 2020)
    if random.random() < 0.5:
        company.size = random.choice(CompanySize.CHOICES)[0]
    if random.random() < 0.5:
        company.type = random.choice(CompanyType.CHOICES)[0]
    if random.random() < 0.5:
        company.organic_products = True
    if random.random() < 0.5:
        company.private_label = True
    if random.random() < 0.5:
        company.female_leadership = True
    if random.random() < 0.5:
        company.branded_value = True
    if random.random() >= 0.1:
        company.is_verified = True
    # company.photo = fake.uri()
    # company.photo_alt = company.name + " Company Photo"

    if save:
        company.save()

    return company


def create_fake_tracking(save=True):
    ip = fake.ipv4()
    ua = fake.user_agent()
    ua_data = parse(ua)
    device_type = get_device_type(ua_data)
    tracking = Tracking(
        ip=ip,
        referrer=random.choice(REFERRER_OPTIONS),
        device_type=device_type,
        device=ua_data.device.family,
        browser=ua_data.browser.family,
        browser_version=ua_data.browser.version_string,
        system=ua_data.os.family,
        system_version=ua_data.os.version_string,
    )

    geo_data = get_geo_details_from_ip(ip)
    if geo_data:
        country = geo_data.get("country")
        region = geo_data.get("region"),
        city = geo_data.get("city")
        postal = geo_data.get("postal")
        location_details = geo_data.get("location_details")
        if location_details:
            if type(location_details) is list:
                location_details = location_details[0]
        tracking.country = country[0] if type(country) is tuple else country
        tracking.region = region[0] if type(region) is tuple else region
        tracking.city = city[0] if type(city) is tuple else city
        tracking.postal = postal[0] if type(postal) is tuple else postal
        tracking.location_details = location_details

    random_value = random.random()

    if random_value <= 0.5:
        tracking.type = TrackingTypes.COMPANY.upper()
        if Company.objects.all():
            tracking.company = Company.objects.all().order_by("?")[0]
    elif 0.5 < random_value <= 0.5:
        tracking.type = TrackingTypes.CATEGORY.upper()
        if Category.objects.all():
            tracking.category = Category.objects.all().order_by("?")[0]
    elif 0.5 < random_value <= 0.75:
        tracking.type = TrackingTypes.PRODUCT.upper()
        if Product.objects.all():
            tracking.product = Product.objects.all().order_by("?")[0]
    else:
        tracking.type = TrackingTypes.OTHER

    if random_value <= 0.5:
        tracking.user = User.objects.accounts().order_by("?")[0]

    if save:
        tracking.save()

    return tracking


def create_account_events(how_many=30):
    for index in range(how_many):
        for i in range(random.randint(6, 12)):
            event = create_fake_event()
            yield "Event type: %s" % (event.type,)


def create_users(how_many=30):
    for dummy in range(how_many):
        user = create_fake_user()
        yield "User: %s" % (user.email,)


def create_users_and_companies(how_many=30):
    for dummy in range(how_many):
        user = create_fake_user()
        company = create_fake_company(user=user)
        yield "User %s and Company %s" % (user.email, company.name)


def create_permission_groups():
    super_users = User.objects.filter(is_superuser=True)
    if not super_users:
        super_users = create_staff_users(1, True)
    group = create_group("Full Access", get_permissions(), super_users)
    yield f"Group: {group}"

    staff_users = create_staff_users()
    customer_support_codenames = [
        perm.codename
        for enum in [ProfilePermissions, OrderPermissions]
        for perm in enum
    ]
    customer_support_codenames.append(AccountPermissions.MANAGE_USERS.codename)
    customer_support_permissions = Permission.objects.filter(
        codename__in=customer_support_codenames
    )
    group = create_group("Customer Support", customer_support_permissions, staff_users)
    yield f"Group: {group}"


def create_news():
    data_news = {
        1: {
            "title": "Public News 1",
            "slug": "public-news-1",
            "summary": "Public News 1 summary",
            "audience": NewsAudienceType.PUBLIC,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Public News 1", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example news for public audience."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        2: {
            "title": "Public News 2",
            "slug": "public-news-2",
            "summary": "Public News 2 summary",
            "audience": NewsAudienceType.PUBLIC,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Public News 2", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example news for public audience."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        3: {
            "title": "Platform News 1",
            "slug": "platform-news-1",
            "summary": "Platform News 1 summary",
            "audience": NewsAudienceType.PLATFORM,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Platform News 1", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example news for platform audience."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        4: {
            "title": "Platform News 2",
            "slug": "platform-news-2",
            "summary": "Platform News 2 summary",
            "audience": NewsAudienceType.PLATFORM,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Platform News 2", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example news for platform audience."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        5: {
            "title": "Staff News 1",
            "slug": "staff-news-1",
            "summary": "Staff News 1 summary",
            "audience": NewsAudienceType.STAFF,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Staff News 1", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example news for staff members."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        6: {
            "title": "Staff News 2",
            "slug": "staff-news-2",
            "summary": "Staff News 2 summary",
            "audience": NewsAudienceType.STAFF,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Staff News 2", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example news for staff members."},
                        "type": "paragraph",
                    },
                ]
            },
        },
    }

    for pk in range(1, 7):
        data = data_news[pk]
        news_data = {
            "content": data["content"],
            "title": data["title"],
            "summary": data["summary"],
            "audience": data["audience"],
            "is_published": True,
        }
        news, _ = News.objects.get_or_create(
            pk=pk, slug=data["slug"], defaults=news_data
        )
        yield "News %s created" % news.slug


def create_companies(how_many=30):
    for dummy in range(how_many):
        company = create_fake_company()
        yield "Company: %s" % (company.name,)


def company_cell_data(reader, row, column):
    return reader.cell(row=row, column=column).value


def company_address_create(**kwargs):
    return Address.objects.create(**kwargs)


def company_user_create(address, **kwargs):
    if User.objects.filter(email=kwargs['email']).exists():
        user = User.objects.get(email=kwargs['email'])
    else:
        user = User.objects.create(**kwargs)
        user.addresses.add(address)
    return user


def company_info_data(reader, row):
    company = {}
    email = company_cell_data(reader, row, 14)
    if email is None:
        email = company_cell_data(reader, row, 1).lower().replace(" ", ".")
        email = email + '@' + email.replace(".", "") + '.com'
    name = company_cell_data(reader, row, 4)
    name = name if name else company_cell_data(reader, row, 1)
    phone_number = company_cell_data(reader, row, 13)
    if type(phone_number) is not str:
        phone_number = None
    address = company_address_create(
        address_name="Office",
        first_name=company_cell_data(reader, row, 1),
        company_name=name,
        street_address_1=company_cell_data(reader, row, 15),
        street_address_2=company_cell_data(reader, row, 16),
        city=company_cell_data(reader, row, 18),
        postal_code=company_cell_data(reader, row, 20),
        country='TR',
        country_area=company_cell_data(reader, row, 17),
        phone=phone_number
    )
    user = company_user_create(
        address=address,
        password=make_password("Company&Password@95164$"),
        email=email,
        phone=phone_number,
        is_seller=True
    )
    founded_year = company_cell_data(reader, row, 11)
    company['user'] = user
    company['name'] = name
    company['slug'] = slugify(name)
    company['website'] = company_cell_data(reader, row, 5)
    company['content_plaintext'] = company_cell_data(reader, row, 7)
    company['founded_year'] = founded_year if isinstance(founded_year, int) else None
    company['no_of_employees'] = company_cell_data(reader, row, 12)
    company['phone'] = phone_number
    company['address'] = address
    return company


def create_sheet_companies():
    wb = openpyxl.load_workbook("./././Koytola Working Sheet.xlsx")
    reader = wb['Şirket Bilgileri']
    # Address.objects.filter(country=None).update(country="TR")

    for i in range(14, 40):
        d = company_info_data(reader, i)
        if not Company.objects.filter(user=d['user'], name=d['name']).exists():
            company_obj = Company.objects.create(**d)
            yield "Company: %s" % (company_obj.name,)


def create_product_categories():
    cats = json.load(open("./././categories.json"))
    for key, value in cats.items():
        if not Category.objects.filter(name=key).exists():
            img_file = open(cats[key]["category_image"], 'rb')
            image = io.BytesIO(img_file.read())
            parent = Category(name=key, slug=slugify(key))
            parent.background_image.save(cats[key]["category_image"].split('/')[1], image)
            parent.save()
        cats[key].pop('category_image')
        if not Category.objects.filter(name="Other", parent=None).exists():
            Category.objects.create(name="Other", slug='other', parent=None)
        parent = Category.objects.get(name=key)
        if not Category.objects.filter(name=f"Other", parent=parent).exists():
            Category.objects.create(name=f"Other", slug=f'other', parent=parent)
        for k, v in value.items():
            keys = {}
            for i in v['keys']:
                keys[i] = ""
            if Category.objects.filter(name=k).exists():
                print("Category: ", k)
                sub = Category.objects.get(name=k)
            else:
                sub = Category.objects.create(name=k, slug=slugify(k), parent=parent, description=keys)
            if not Category.objects.filter(name=f"Other", parent=sub).exists():
                Category.objects.create(name=f"Other", slug=f'other', parent=sub)
            for cat in v['category']:
                if Category.objects.filter(slug=slugify(cat)).exists():
                    slug = slugify(k) + '-' + slugify(cat)
                else:
                    slug = slugify(cat)
                if not Category.objects.filter(name=cat).exists():
                    Category(name=cat, slug=slug, parent=sub).save()
                    yield "Category: %s" % (cat,)


def create_staffs():
    for permission in get_permissions():
        base_name = permission.codename.split("_")[1:]

        group_name = " ".join(base_name)
        group_name += " management"
        group_name = group_name.capitalize()

        email_base_name = [name[:-1] if name[-1] == "s" else name for name in base_name]
        user_email = ".".join(email_base_name)
        user_email += ".manager@example.com"

        user = _create_staff_user(email=user_email)
        group = create_group(group_name, [permission], [user])

        yield f"Group: {group}"
        yield f"User: {user}"


def create_tracking(how_many=30):
    for dummy in range(how_many):
        tracking = create_fake_tracking()
        yield "Tracking: %s" % (tracking.ip,)


def create_group(name, permissions, users):
    group, _ = Group.objects.get_or_create(name=name)
    group.permissions.add(*permissions)
    group.user_set.add(*users)
    return group


def _create_staff_user(email=None, superuser=False):
    user = User.objects.filter(email=email).first()
    if user:
        return user
    address = create_address()
    first_name = address.first_name
    last_name = address.last_name
    if not email:
        email = get_email(first_name, last_name)

    staff_user = User.objects.create_user(
        first_name=first_name,
        last_name=last_name,
        email=email,
        password=DUMMY_STAFF_PASSWORD,
        default_billing_address=address,
        default_shipping_address=address,
        is_staff=True,
        is_active=True,
        is_superuser=superuser,
    )
    return staff_user


def create_staff_users(how_many=2, superuser=False):
    users = []
    for _ in range(how_many):
        staff_user = _create_staff_user(superuser=superuser)
        users.append(staff_user)
    return users


def add_address_to_admin(email):
    address = create_address()
    user = User.objects.get(email=email)
    store_user_address(user, address, AddressType.BILLING)
    store_user_address(user, address, AddressType.SHIPPING)


def create_page_types():
    data = [
        {
            "pk": 1,
            "fields": {
                "private_metadata": {},
                "metadata": {},
                "name": "About",
                "slug": "about",
            },
        },
        {
            "pk": 2,
            "fields": {
                "private_metadata": {},
                "metadata": {},
                "name": "Blog",
                "slug": "blog",
            },
        },
        {
            "pk": 3,
            "fields": {
                "private_metadata": {},
                "metadata": {},
                "name": "Company",
                "slug": "company-details",
            },
        },
    ]
    for page_type_data in data:
        pk = page_type_data.pop("pk")
        page_type, _ = PageType.objects.update_or_create(
            pk=pk, **page_type_data["fields"]
        )
        yield "Page type %s created" % page_type.slug


def create_hscode_and_product():
    # HSCodeAndProduct.objects.all().delete()
    reader = openpyxl.load_workbook("./././HS code List.xlsx").active
    for i in range(2, reader.max_row):
        if not HSCodeAndProduct.objects.filter(hs_code=reader.cell(row=i, column=1).value).exists():
            HSCodeAndProduct(
                hs_code=reader.cell(row=i, column=1).value,
                product_name=reader.cell(row=i, column=2).value
            ).save()
        yield "HS Code And Product Created %s" % reader.cell(row=i, column=1).value


def create_pages():
    data_pages = {
        1: {
            "title": "About",
            "slug": "about",
            "summary": "About example page for the project",
            "page_type_id": 1,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Trade-Tech for the PWA era", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {
                            "text": (
                                "A modular, high performance trade marketplace "
                                "built with GraphQL, Django, and ReactJS."
                            ),
                            "level": 2,
                        },
                        "type": "header",
                    },
                    {"data": {"text": ""}, "type": "paragraph"},
                    {
                        "data": {
                            "text": (
                                "Koytola is a rapidly-growing trade-tech marketplace "
                                "platform that has served high-volume companies "
                                "from branches like posting and publishing since 2021. "
                                "Based on Python and Django, the latest major update "
                                "introduces a modular front end with a GraphQL API "
                                "and site-front and dashboard written in React "
                                "to make Project a full-functionality "
                                "open source platform."
                            )
                        },
                        "type": "paragraph",
                    },
                    {"data": {"text": ""}, "type": "paragraph"},
                    {
                        "data": {
                            "text": (
                                '<a href="https://koytola.com">'
                                "Contact us today!</a>"
                            )
                        },
                        "type": "paragraph",
                    },
                ],
            },
        },
        2: {
            "title": "Blog Home",
            "slug": "blogs",
            "summary": "Blog home example page for the project",
            "page_type_id": 2,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Blogs", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example blogs home page."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        3: {
            "title": "Profile Guidelines",
            "slug": "profile-guidelines",
            "summary": "Company guidelines example page for the project",
            "page_type_id": 3,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Profile Guidelines", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example company profile guidelines page."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        4: {
            "title": "Example Page 1",
            "slug": "example-page-1",
            "summary": "Example page 1 for the project",
            "page_type_id": 3,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Example Page 1", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example page."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        5: {
            "title": "Example Page 2",
            "slug": "example-page-2",
            "summary": "Example page 2 for the project",
            "page_type_id": 3,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Example Page 2", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example page."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        6: {
            "title": "Blog 1",
            "slug": "blog-1",
            "summary": "Blog page 1 for the project",
            "page_type_id": 2,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Blog Page 1", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example blog page."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        7: {
            "title": "Blog 2",
            "slug": "blog-2",
            "summary": "Blog page 2 for the project",
            "page_type_id": 2,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Blog Page 2", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example blog page."},
                        "type": "paragraph",
                    },
                ]
            },
        },
        8: {
            "title": "Blog 3",
            "slug": "blog-3",
            "summary": "Blog page 3 for the project",
            "page_type_id": 2,
            "content": {
                "blocks": [
                    {
                        "data": {"text": "Blog Page 3", "level": 2},
                        "type": "header",
                    },
                    {
                        "data": {"text": "This is example blog page."},
                        "type": "paragraph",
                    },
                ]
            },
        },
    }

    for pk in range(1, 9):
        data = data_pages[pk]
        page_data = {
            "content": data["content"],
            "title": data["title"],
            "summary": data["summary"],
            "is_published": True,
            "page_type_id": data["page_type_id"],
        }
        page, _ = Page.objects.get_or_create(
            pk=pk, slug=data["slug"], defaults=page_data
        )
        yield "Page %s created" % page.slug


def generate_menu_items(
        menu: Menu,
        parent_menu_item,
        page: Page = None,
        url: str = None,
        url_title: str = None,
):
    if page:
        menu_item, created = menu.items.get_or_create(
            name=page.title, page=page, parent=parent_menu_item
        )
        if created:
            yield "Created menu item for page %s" % page

    elif url:
        menu_item, created = menu.items.get_or_create(
            name=url_title, url=url, parent=parent_menu_item
        )
        if created:
            yield "Created menu item for url %s" % url_title


def generate_menu_tree(menu):
    page_types = PageType.objects.all()

    generate_menu_items(menu, "/", None)

    for page_type in page_types:
        pages = Page.objects.filter(page_type=page_type)
        for page in pages:
            for msg in generate_menu_items(menu, page, None):
                yield msg


def create_menus():
    # Create navbar menu
    top_menu, _ = Menu.objects.get_or_create(
        name=settings.DEFAULT_MENUS["top_menu_name"],
    )
    top_menu.items.all().delete()
    page_type_1 = PageType.objects.all().order_by("?")[0]
    item, _ = top_menu.items.get_or_create(
        name=page_type_1.name, url=page_type_1.slug
    )
    item_site = top_menu.items.get_or_create(name="Home", url="/")[0]
    item_site = top_menu.items.get_or_create(name="Services", url="/services/")[0]
    item_site = top_menu.items.get_or_create(name="Team", url="/team/")[0]

    page = Page.objects.order_by("?")[0]
    item_site.children.get_or_create(name=page.title, page=page, menu=top_menu)
    yield "Created navbar menu"
    for msg in generate_menu_tree(top_menu):
        yield msg

    # Create footer menu with page types
    bottom_menu, _ = Menu.objects.get_or_create(
        name=settings.DEFAULT_MENUS["bottom_menu_name"]
    )
    bottom_menu.items.all().delete()
    page_type_2 = PageType.objects.all().order_by("?")[0]
    item, _ = bottom_menu.items.get_or_create(
        name=page_type_2.name, url=page_type_2.slug
    )
    item_site = bottom_menu.items.get_or_create(name="Project", url="/")[0]

    page = Page.objects.order_by("?")[0]
    item_site.children.get_or_create(name=page.title, page=page, menu=bottom_menu)

    api_url = build_absolute_uri(reverse("api"))
    item_site.children.get_or_create(
        name="GraphQL API", url=api_url, menu=bottom_menu
    )

    yield "Created footer menu"
    site = Site.objects.get_current()
    site_settings = site.settings
    site_settings.top_menu = top_menu
    site_settings.bottom_menu = bottom_menu
    site_settings.save()


def get_image(image_dir, image_name):
    img_path = os.path.join(image_dir, image_name)
    return File(open(img_path, "rb"), name=image_name)


def create_countries():
    csv_reader = list(csv.reader(open('./././countries.csv')))
    for i in csv_reader[1:]:
        if not Countries.objects.filter(code=i[0]).exists():
            Countries(code=i[0], name=i[3], latitude=i[1], longitude=i[2]).save()
        yield "country %s" % i[3]


def create_countries_flag():
    csv_reader = list(csv.reader(open('./././Country_Flags.csv')))[1:]
    for i in csv_reader:
        if i[2] and Countries.objects.filter(name=i[0], flag=None).exists():
            try:
                response = urllib.request.urlopen(i[2])
                if response.getcode() == 200:
                    obj = io.BytesIO(response.read())
                    company = Countries.objects.get(name=i[0])
                    company.flag.save(i[1], obj)
                    company.save()
                yield "country flag %s" % i[0]
            except urllib.error.HTTPError:
                continue


def rosetter_save(i, type, category):
    obj = Roetter(type=type, name=i['name'], category=category)
    img_file = open('./././Rosette/' + i['image'], 'rb')
    image = io.BytesIO(img_file.read())
    obj.image.save(i['image'], image)
    obj.save()
    return obj


def create_rosetter():
    rosetters = json.load(open('./././rosetters.json', 'r'))
    for i in rosetters['Company']:
        if not Roetter.objects.filter(name=i['name'], type='Company').exists():
            company = rosetter_save(i, 'Company', None)
        yield f"company rosetter {i['name'] if 'name' in i else ''}"
    for i in rosetters['Product']:
        for j in rosetters['Product'][i]['rosetter']:
            if not Roetter.objects.filter(name=j['name'], type='Product').exists():
                one = rosetter_save(j, 'Product', rosetters['Product'][i]['category'])
            yield f"Product rosetter {j['name'] if 'name' in j else ''}"


def create_industry():
    industry = ['Cereals, Pulses, Oil Seeds and Products', 'Food and Beverage',
                'Dried Fruit', 'Hazelnut and Products', 'Olive and Olive Oil Products', 'Tobacco',
                'Ornamental Plants and Products', 'Aqua and Animal Products', 'Wood and Forestry Products',
                'Textile and Raw Materials', 'Leather and Leather products', 'Carpet',
                'Chemicals and chemical products', 'Apparel', 'Automotive', 'Ship and Yacht',
                'Electric, Electronic and Service', 'Machinery and Machinery Accessories', 'Other',
                'Ferrous and Non - Ferrous Metals', 'Steel', 'Cement, Glass, Ceramic and Soil Products', 'Jewelry',
                'Defense and Aerospace', 'Other Industry Products', 'Mining Products', 'Home Appliance', 'Furniture',
                'Cosmetics', 'Gift', 'Pet Food and Products', 'Fresh Fruit and Vegetable', 'Sport and Hobby Equipment'
                ]
    for i in industry:
        if not Industry.objects.filter(name=i).exists():
            Industry(name=i).save()
            yield "Industry %s" % i


def create_product_certificate():
    certificate = ['ACRS', 'AENOR', 'AGA', 'ANVISA', 'API', 'AQSIQ', 'ASME', 'BASEC', 'BGIA', 'BHRC', 'CEDD', 'CERTIF',
                   'CFDA', 'CQC', 'CSTB', 'CSIC', 'DTI', 'DHI', 'DVGW', 'EASA', 'EPA', 'EPAL', 'ESMA', 'EUCEB', 'OCB',
                   'FAA-L', 'FDA', 'FSEC', 'GAZPROM', 'ICC', 'ICONTEC', 'ISI', 'ITB', 'KEBS', 'KONTROLLRADET', 'NOM',
                   'THE CHIEF RABBINATE OF ISRAEL**', 'LEATHER WORKING GROUP', 'LEED', 'LGA', 'LPCB', 'NAFDAC', 'NEMKO',
                   'POLITECNICO DI TORINO (DM 17.01.2018)', 'PSQCA', 'ROSSELKHOZNADZOR', 'ROSZDRAVNADZOR', 'SASO',
                   'SEMKO', 'SFDA', 'SII', 'SIRIM QAS', 'SNI', 'SON', 'SRCC', 'TSE', 'USDA', 'VCA', "ZDHC",
                   'THE FRIEND OF THE SEA CERTIFICATION', 'THE VEGAN SOCIETY CERTIFICATE OF REGISTRATION',
                   'ACS', 'AS / EN / JISQ 9100', 'AD 200 CODE', 'ATEX', 'BASEC', 'BRC', 'CARB', 'CD', 'CCC', 'CFDA',
                   'CQC', 'CE', 'CEKAL', 'CLASS', 'CMIM MARK', 'CMMI DEV', 'Crash Test Certificate', 'CTUV', 'ECE',
                   'EC Type Examination', 'ECM', 'ECOLABEL', 'ECOCERT', 'EMC', 'EN', 'ETA', 'EUROVENT', 'FDA',
                   'F4 STAR', 'FIBC', 'Florida Product Approval', 'FM Approvals', 'FSSC 22000', 'G MARK', 'GASTEC QA',
                   'GMP', 'GMP + FSA', 'GOST', 'GS', 'HACCP', 'HAR', 'HALAL', 'INSTA - CERT', 'ISCC', 'ISO',
                   'KEMA KEUR', 'KEYMARK', 'KITEMARK', 'KOSHER', 'KRAV', 'Leather Working Group', 'Watermark',
                   'Manufacturing Approval of Steel Forging', 'Master Certificate', 'Marine Certificate', 'MDR',
                   'Miami Dade County Notice', 'National Technical Assessment', 'NF', 'Organic Certificates', 'Q Mark',
                   'RAL', 'Recycled Claim Standard', 'RTN', 'SIL', 'S Mark', 'Solar Certificate', 'Standardsmark',
                   'Toxproof', 'TSE', 'TSI', 'UKCA', 'USDA', 'Vehicle Spare Parts', 'Vegetarian Certificate', 'WRAS'
                   ]
    for i in certificate:
        if not CertificateType.objects.filter(type="Product", name=i).exists():
            obj = CertificateType.objects.create(type="Product", name=i)
        else:
            obj = CertificateType.objects.get(type="Product", name=i)
        image_name = obj.name[0].upper()+'.png'
        img_file = open('./././certificate_icon/' + image_name, 'rb')
        image = io.BytesIO(img_file.read())
        obj.image.save(image_name, image)
        obj.save()
        yield "Product certificate %s" % i


def create_company_certificate():
    certificate = ['ACS', 'ADR', 'AD 2000 CODE', 'AS/EN/JISQ 9100', 'ATEX', 'BASEC', 'BRC', 'CARB', 'CB', 'CCC', 'CE',
                   'CEKAL', 'CLASS', 'CMIM MARK', 'CMMI DEV', 'CRASH TEST CERTIFICAT', 'CTUV', 'DIN', 'ECE', 'EC',
                   'ECM', 'ECOLABEL', 'ECOCERT', 'EMC', 'EN', 'ETA', 'EUROVENT', 'F4 STAR', 'FIBC', 'KEMA KEUR',
                   'FLORIDA PRODUCT APPROVAL', 'FM APPROVALS', 'FSSC 22000', 'G MARK', 'GASTEC QA', 'GMP', 'GMP+ FSA',
                   'GOST', 'GS', 'HACCP', 'HAR', 'HELAL', 'INSTA-CERT', 'HİJYEN', 'ISCC', "MASTER CERTIFICATE",
                   'ISO 13485*, 14001*, 14064*, 17025*, 22000*, 27001*, 50001*', 'İYİ TARIM UYGULAMALARI', 'KEYMARK KITEMARK',
                   'KRAV CERTIFICATION', 'MANUFACTURING APPROVAL OF STEEL FORGING', 'MARINE CERTIFICATE',
                   'MDR', 'MIAMI DADE COUNTY NOTICE OF ACCEPTANCE', 'MIL-STD-461F*', 'MIL-STD-810G*',
                   'NATIONAL TECHNICAL ASSESMENT', 'VI-1', 'WRAS', 'WATERMARK', 'YANGIN', 'ACS', 'AS / EN / JISQ 9100',
                   'AD 200 CODE', 'ATEX', 'BASEC', 'BRC', 'CARB', 'CD', 'CCC', 'CFDA',
                   'CQC', 'CE', 'CEKAL', 'CLASS', 'CMIM MARK', 'CMMI DEV', 'Crash Test Certificate', 'CTUV', 'ECE',
                   'EC Type Examination', 'ECM', 'ECOLABEL', 'ECOCERT', 'EMC', 'EN', 'ETA', 'EUROVENT', 'FDA',
                   'F4 STAR', 'FIBC', 'Florida Product Approval', 'FM Approvals', 'FSSC 22000', 'G MARK', 'GASTEC QA',
                   'GMP', 'GMP + FSA', 'GOST', 'GS', 'HACCP', 'HAR', 'HALAL', 'INSTA - CERT', 'ISCC', 'ISO',
                   'KEMA KEUR', 'KEYMARK', 'KITEMARK', 'KOSHER', 'KRAV', 'Leather Working Group', 'Watermark',
                   'Manufacturing Approval of Steel Forging', 'Master Certificate', 'Marine Certificate', 'MDR',
                   'Miami Dade County Notice', 'National Technical Assessment', 'NF', 'Organic Certificates', 'Q Mark',
                   'RAL', 'Recycled Claim Standard', 'RTN', 'SIL', 'S Mark', 'Solar Certificate', 'Standardsmark',
                   'Toxproof', 'TSE', 'TSI', 'UKCA', 'USDA', 'Vehicle Spare Parts', 'Vegetarian Certificate', 'WRAS'
                   ]
    for i in certificate:
        if not CertificateType.objects.filter(type="Company", name=i).exists():
            obj = CertificateType.objects.create(type="Company", name=i)
        else:
            obj = CertificateType.objects.get(type="Company", name=i)
        image_name = 'I.png' if obj.name[0] == "İ" else obj.name[0].upper() + '.png'
        img_file = open('./././certificate_icon/' + image_name, 'rb')
        image = io.BytesIO(img_file.read())
        obj.image.save(image_name, image)
        obj.save()
        yield "Company certificate %s" % i


def google_news_create():
    url = '''https://news.google.com/rss/search?q=trade+site:ft.com+OR+site:bloomberg.com+OR+site:forbes.com+OR+site:theguardian.com&client=safari&rls=en&hl=en-US&gl=US&ceid=US:en'''
    f = requests.get(url)
    news = xmltodict.parse(f.content.decode('utf-8'))
    for i in news['rss']['channel']['item']:
        pub_date = datetime.strptime(i['pubDate'], "%a, %d %b %Y %H:%M:%S %Z")
        if not News.objects.filter(slug=slugify(i['title'])).exists():
            News(title=i['title'], slug=slugify(i['title']), link=i['link'], publication_date=pub_date).save()
